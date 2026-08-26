"""Archive safety, data quality profiling and document batch orchestration."""

from __future__ import annotations

from collections import Counter, defaultdict
import csv
from datetime import datetime
import hashlib
import json
import math
from pathlib import Path, PurePosixPath
import shutil
import statistics
import tarfile
from typing import Any, Iterable
import zipfile

from ..artifacts import artifact, safe_output_path, sha256_file, write_report
from ..capabilities import CapabilityResolver
from ..models import SkillStatus, invalid, result


def _unsafe_member(name: str) -> bool:
    normalized = name.replace("\\", "/")
    path = PurePosixPath(normalized)
    return path.is_absolute() or ".." in path.parts or (len(normalized) > 1 and normalized[1] == ":")


def _archive_entries(path: Path):
    entries = []
    if zipfile.is_zipfile(path):
        with zipfile.ZipFile(path) as package:
            for info in package.infolist():
                if info.is_dir(): continue
                with package.open(info) as stream:
                    digest = hashlib.sha256(stream.read()).hexdigest()
                entries.append({"name": info.filename, "size": info.file_size, "compressed_size": info.compress_size, "checksum": digest, "unsafe": _unsafe_member(info.filename), "hidden": any(part.startswith(".") for part in PurePosixPath(info.filename).parts), "nested_archive": Path(info.filename).suffix.lower() in {".zip", ".7z", ".tar", ".tgz", ".gz"}})
        return "zip", entries
    if tarfile.is_tarfile(path):
        with tarfile.open(path) as package:
            for info in package.getmembers():
                if not info.isfile(): continue
                stream = package.extractfile(info)
                digest = hashlib.sha256(stream.read()).hexdigest() if stream else None
                entries.append({"name": info.name, "size": info.size, "compressed_size": None, "checksum": digest, "unsafe": _unsafe_member(info.name) or info.issym() or info.islnk(), "hidden": any(part.startswith(".") for part in PurePosixPath(info.name).parts), "nested_archive": Path(info.name).suffix.lower() in {".zip", ".7z", ".tar", ".tgz", ".gz"}})
        return "tar", entries
    if path.suffix.lower() == ".7z":
        return "7z", None
    raise ValueError("Unsupported or invalid archive")


def _tree(names: Iterable[str]) -> list[str]:
    output = []
    for name in sorted(names):
        parts = PurePosixPath(name).parts
        output.append("  " * max(0, len(parts) - 1) + ("└─ " if len(parts) > 1 else "") + parts[-1])
    return output


def _archive(request: dict[str, Any]) -> dict[str, Any]:
    path = Path(str(request.get("input") or ""))
    if not path.is_file():
        return invalid("input must reference an existing archive")
    operation = str(request.get("operation") or "inspect")
    try:
        kind, entries = _archive_entries(path)
    except Exception as exc:
        return result(SkillStatus.INVALID_INPUT, "Archive cannot be inspected", error_code=type(exc).__name__, error_detail=str(exc))
    if kind == "7z" and entries is None:
        capability = CapabilityResolver().resolve("7z")
        return result(SkillStatus.DEPENDENCY_MISSING, "7z CLI support is not available", error_code="SEVEN_ZIP_MISSING", capabilities={"7z": capability})
    assert entries is not None
    unsafe = [item["name"] for item in entries if item["unsafe"]]
    duplicate_groups = defaultdict(list)
    for item in entries: duplicate_groups[item["checksum"]].append(item["name"])
    duplicates = {key: value for key, value in duplicate_groups.items() if key and len(value) > 1}
    if operation == "compare":
        other = Path(str(request.get("other") or ""))
        if not other.is_file(): return invalid("compare requires other archive")
        try: _other_kind, other_entries = _archive_entries(other)
        except Exception as exc: return invalid(f"other archive cannot be inspected: {exc}")
        if other_entries is None: return result(SkillStatus.DEPENDENCY_MISSING, "7z CLI support is required", error_code="SEVEN_ZIP_MISSING")
        left = {item["name"]: item["checksum"] for item in entries}; right = {item["name"]: item["checksum"] for item in other_entries}
        data = {"added": sorted(set(right) - set(left)), "removed": sorted(set(left) - set(right)), "changed": sorted(name for name in set(left) & set(right) if left[name] != right[name])}
        return result(SkillStatus.SUCCESS, "Archives compared without extraction", data=data)
    if operation == "extract":
        if unsafe:
            return result(SkillStatus.INVALID_INPUT, "Extraction blocked because unsafe paths were detected", data={"unsafe": unsafe}, error_code="PATH_TRAVERSAL")
        output_root = safe_output_path(request, source=path, stem_suffix="extracted", extension=".directory").with_suffix("")
        output_root.mkdir(parents=True, exist_ok=False)
        if kind == "zip":
            with zipfile.ZipFile(path) as package:
                for info in package.infolist():
                    if info.is_dir(): continue
                    destination = output_root.joinpath(*PurePosixPath(info.filename).parts)
                    destination.parent.mkdir(parents=True, exist_ok=True)
                    with package.open(info) as source_stream, destination.open("wb") as output_stream: shutil.copyfileobj(source_stream, output_stream)
        else:
            with tarfile.open(path) as package:
                for info in package.getmembers():
                    if not info.isfile(): continue
                    destination = output_root.joinpath(*PurePosixPath(info.name).parts)
                    destination.parent.mkdir(parents=True, exist_ok=True)
                    source_stream = package.extractfile(info)
                    if source_stream:
                        with source_stream, destination.open("wb") as output_stream: shutil.copyfileobj(source_stream, output_stream)
        manifest = output_root / "EXTRACTION_MANIFEST.json"
        manifest.write_text(json.dumps({"source": path.name, "files": [item["name"] for item in entries]}, indent=2), encoding="utf-8")
        item = artifact(manifest, operation="safe_extract", source=path)
        return result(SkillStatus.SUCCESS, "Archive safely extracted to a new directory", data={"directory": output_root.name, "files": len(entries)}, artifacts=[item])
    classifications = Counter((Path(item["name"]).suffix.lower() or "no_extension") for item in entries)
    sections = [("Summary", [f"format={kind}", f"files={len(entries)}", f"uncompressed_size={sum(item['size'] for item in entries)}"]), ("Directory Tree", _tree(item["name"] for item in entries)), ("Security", unsafe or ["No path traversal entry detected"]), ("Hidden Files", [item["name"] for item in entries if item["hidden"]]), ("Nested Archives", [item["name"] for item in entries if item["nested_archive"]])]
    item = _markdown(request, path, "Archive Inspection", sections, "archive-inspector")
    return result(SkillStatus.SUCCESS, "Archive inspected without extraction", data={"format": kind, "file_count": len(entries), "total_size": sum(item["size"] for item in entries), "classifications": dict(classifications), "unsafe_paths": unsafe, "duplicates": duplicates, "entries": entries}, artifacts=[item])


def _markdown(request, source: Path | None, title: str, sections, operation: str):
    lines = [f"# {title}", ""]
    for heading, values in sections:
        lines += [f"## {heading}", ""]
        if isinstance(values, list): lines += [f"- {value}" for value in values] or ["- None"]
        else: lines.append(str(values))
        lines.append("")
    _path, item = write_report(request, "\n".join(lines), operation=operation, source=source)
    return item


def _load_tabular(path: Path):
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open(encoding="utf-8-sig", errors="replace", newline="") as stream:
            rows = list(csv.DictReader(stream, delimiter=delimiter))
        return list(rows[0].keys()) if rows else [], rows, suffix.lstrip(".")
    if suffix == ".jsonl":
        rows = [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
        columns = sorted({key for row in rows if isinstance(row, dict) for key in row})
        return columns, rows, "jsonl"
    if suffix == ".xlsx":
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True); sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True); header = next(iterator, ())
        columns = [str(value or f"column_{index + 1}") for index, value in enumerate(header)]
        rows = [dict(zip(columns, values)) for values in iterator]
        workbook.close(); return columns, rows, "xlsx"
    if suffix == ".parquet":
        import pyarrow.parquet as pq
        table = pq.read_table(path); rows = table.to_pylist(); return table.column_names, rows, "parquet"
    raise ValueError("Unsupported data format")


def _numeric(values):
    output = []
    for value in values:
        if value in (None, ""): continue
        try: output.append(float(value))
        except (ValueError, TypeError): return None
    return output


def _percentile(values, percent):
    if not values: return None
    ordered = sorted(values); position = (len(ordered) - 1) * percent; lower = math.floor(position); upper = math.ceil(position)
    if lower == upper: return ordered[lower]
    return ordered[lower] * (upper - position) + ordered[upper] * (position - lower)


def _profile(request: dict[str, Any]) -> dict[str, Any]:
    path = Path(str(request.get("input") or ""))
    if not path.is_file(): return invalid("input must reference CSV, TSV, XLSX, JSONL or Parquet")
    capabilities = CapabilityResolver().resolve_many(("openpyxl", "pyarrow"))
    if path.suffix.lower() == ".xlsx" and not capabilities["openpyxl"]["available"]:
        return result(SkillStatus.DEPENDENCY_MISSING, "openpyxl is required for XLSX profiling", error_code="OPENPYXL_MISSING", capabilities=capabilities)
    if path.suffix.lower() == ".parquet" and not capabilities["pyarrow"]["available"]:
        return result(SkillStatus.DEPENDENCY_MISSING, "pyarrow is required for Parquet profiling", error_code="PYARROW_MISSING", capabilities=capabilities)
    try: columns, rows, data_format = _load_tabular(path)
    except Exception as exc: return result(SkillStatus.INVALID_INPUT, "Dataset cannot be parsed", error_code=type(exc).__name__, error_detail=str(exc), capabilities=capabilities)
    row_hashes = [json.dumps(row, ensure_ascii=False, sort_keys=True, default=str) for row in rows]
    duplicate_rows = len(row_hashes) - len(set(row_hashes))
    profiles = {}
    numeric_columns = {}
    for column in columns:
        values = [row.get(column) for row in rows]
        nulls = sum(value is None or value == "" for value in values)
        non_null = [value for value in values if value not in (None, "")]
        counts = Counter(str(value) for value in non_null)
        numbers = _numeric(non_null)
        profile = {"nulls": nulls, "unique": len(counts), "constant": len(counts) <= 1 and bool(non_null), "top_values": counts.most_common(10)}
        if numbers is not None and numbers:
            q1 = _percentile(numbers, .25); q3 = _percentile(numbers, .75); iqr = (q3 - q1) if q1 is not None and q3 is not None else 0
            profile.update({"type": "numeric", "min": min(numbers), "max": max(numbers), "mean": statistics.fmean(numbers), "median": statistics.median(numbers), "std": statistics.pstdev(numbers) if len(numbers) > 1 else 0, "percentiles": {"p25": q1, "p50": _percentile(numbers, .5), "p75": q3, "p95": _percentile(numbers, .95)}, "outliers": sum(value < q1 - 1.5 * iqr or value > q3 + 1.5 * iqr for value in numbers) if iqr else 0})
            numeric_columns[column] = numbers
        else: profile["type"] = "categorical"
        if profile["type"] == "categorical" and non_null:
            parsed_times = []
            for value in non_null:
                try: parsed_times.append(datetime.fromisoformat(str(value).replace("Z", "+00:00")))
                except ValueError: pass
            if len(parsed_times) >= max(2, round(len(non_null) * 0.8)):
                ordered_times = sorted(parsed_times)
                gaps = [(ordered_times[index] - ordered_times[index - 1]).total_seconds() for index in range(1, len(ordered_times))]
                profile.update({"type": "timestamp", "min": ordered_times[0].isoformat(), "max": ordered_times[-1].isoformat(), "max_gap_seconds": max(gaps) if gaps else 0, "median_gap_seconds": statistics.median(gaps) if gaps else 0})
        profiles[column] = profile
    correlations = {}
    for left_index, left in enumerate(numeric_columns):
        for right in list(numeric_columns)[left_index + 1:]:
            pairs = []
            for row in rows:
                try: pairs.append((float(row[left]), float(row[right])))
                except (ValueError, TypeError): pass
            if len(pairs) > 1:
                xs, ys = zip(*pairs); mx, my = statistics.fmean(xs), statistics.fmean(ys); denominator = math.sqrt(sum((x-mx)**2 for x in xs) * sum((y-my)**2 for y in ys))
                correlations[f"{left}~{right}"] = sum((x-mx)*(y-my) for x,y in pairs) / denominator if denominator else None
    label = request.get("label_column"); imbalance = None
    if label in columns:
        counts = Counter(str(row.get(label)) for row in rows); imbalance = {"counts": dict(counts), "ratio": max(counts.values()) / min(counts.values()) if counts and min(counts.values()) else None}
    leakage = []
    split_column = request.get("split_column")
    if split_column in columns:
        seen = defaultdict(set)
        for row, signature in zip(rows, row_hashes): seen[signature].add(str(row.get(split_column)))
        leakage = [signature[:80] for signature, splits in seen.items() if len(splits) > 1]
    summary = {"format": data_format, "rows": len(rows), "columns": len(columns), "schema": columns, "duplicate_rows": duplicate_rows, "profiles": profiles, "correlations": correlations, "class_imbalance": imbalance, "leakage_suspicion": leakage}
    sections = [("Shape", [f"rows={len(rows)}", f"columns={len(columns)}"]), ("Data Quality", [f"duplicate_rows={duplicate_rows}", f"constant_columns={[name for name,value in profiles.items() if value['constant']]}", f"leakage_suspicion={len(leakage)}"]), ("Schema", [f"{name}: {profiles[name]['type']}, nulls={profiles[name]['nulls']}, unique={profiles[name]['unique']}" for name in columns])]
    md_item = _markdown(request, path, "Data Profile", sections, "data-profiler")
    json_output = safe_output_path(request, source=path, stem_suffix="profile", extension=".json"); json_output.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    json_item = artifact(json_output, operation="data-profiler", source=path, extra={"rows": len(rows), "columns": len(columns)})
    artifacts = [md_item, json_item]
    if request.get("xlsx_report"):
        if not capabilities["openpyxl"]["available"]:
            return result(SkillStatus.PARTIAL_SUCCESS, "Markdown/JSON profile generated; XLSX report requires openpyxl", data=summary, artifacts=artifacts, error_code="OPENPYXL_MISSING", capabilities=capabilities)
        from openpyxl import Workbook
        workbook = Workbook(); sheet = workbook.active; sheet.title = "Profile"
        sheet.append(["column", "type", "nulls", "unique", "constant", "min", "max", "mean", "median", "std"])
        for name in columns:
            value = profiles[name]
            sheet.append([name, value.get("type"), value.get("nulls"), value.get("unique"), value.get("constant"), value.get("min"), value.get("max"), value.get("mean"), value.get("median"), value.get("std")])
        xlsx_output = safe_output_path(request, source=path, stem_suffix="profile", extension=".xlsx"); workbook.save(xlsx_output)
        artifacts.append(artifact(xlsx_output, operation="data-profiler", source=path, extra={"rows": len(rows), "columns": len(columns)}))
    return result(SkillStatus.SUCCESS, "Dataset quality profile generated", data=summary, artifacts=artifacts, capabilities=capabilities)


OFFICE_BUILTINS = {".pdf": "pdf", ".docx": "docx", ".xlsx": "xlsx", ".pptx": "pptx"}


def _documents(request: dict[str, Any]) -> dict[str, Any]:
    input_value = request.get("input")
    if not input_value:
        return invalid("input must reference an existing directory")
    root = Path(str(input_value))
    if not root.is_dir(): return invalid("input must reference an existing directory")
    recursive = bool(request.get("recursive", True))
    files = sorted(path for path in (root.rglob("*") if recursive else root.glob("*")) if path.is_file())
    records = []
    hashes = defaultdict(list)
    for path in files:
        digest = sha256_file(path); hashes[digest].append(path.name)
        records.append({"name": path.name, "relative": path.relative_to(root).as_posix(), "extension": path.suffix.lower(), "size": path.stat().st_size, "modified": datetime.fromtimestamp(path.stat().st_mtime).isoformat(), "checksum": digest, "delegate_skill": OFFICE_BUILTINS.get(path.suffix.lower())})
    duplicates = {digest: names for digest, names in hashes.items() if len(names) > 1}
    rules = request.get("rules") or {}
    rename = request.get("rename") or {}
    prefix = str(rename.get("prefix") or "") if isinstance(rename, dict) else ""
    suffix = str(rename.get("suffix") or "") if isinstance(rename, dict) else ""
    plan = []
    for record in records:
        category = str(rules.get(record["extension"], record["extension"].lstrip(".") or "other"))
        original = Path(record["name"])
        renamed = f"{prefix}{original.stem}{suffix}{original.suffix}"
        plan.append({"source": record["relative"], "category": category, "target": f"{category}/{renamed}", "delegate_skill": record["delegate_skill"]})
    applied = []
    applied_artifacts = []
    if request.get("apply"):
        output_value = request.get("output_dir")
        if not output_value:
            return invalid("apply requires output_dir")
        output_dir = Path(str(output_value))
        action = str(request.get("action") or "copy")
        for item in plan:
            source = root / PurePosixPath(item["source"])
            target = output_dir / PurePosixPath(item["target"])
            target.parent.mkdir(parents=True, exist_ok=True)
            if target.exists(): continue
            if action == "move": shutil.move(str(source), str(target))
            else: shutil.copy2(source, target)
            applied.append(item["target"])
            applied_artifacts.append(artifact(target, operation=f"document-batch-{action}", source=source))
    manifest = {"root": root.name, "file_count": len(records), "files": records, "duplicates": duplicates, "plan": plan, "applied": applied, "office_operations": "delegate to QwenPaw builtin pdf/docx/xlsx/pptx Skills"}
    output = safe_output_path(request, source=None, stem_suffix="document-index", extension=".json"); output.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    item = artifact(output, operation="document-batch-processor", extra={"source": root.name})
    return result(SkillStatus.SUCCESS, "Document batch plan generated" + (" and applied" if applied else " in dry-run mode"), data=manifest, artifacts=[item, *applied_artifacts])


def execute(skill_name: str, request: dict[str, Any]) -> dict[str, Any]:
    handlers = {"archive-inspector": _archive, "data-profiler": _profile, "document-batch-processor": _documents}
    handler = handlers.get(skill_name)
    return handler(request) if handler else invalid(f"Unsupported data/file Skill: {skill_name}")
